import json
import re
import sqlite3
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
FRONTEND_DATA = ROOT / "vue-waternet" / "public" / "data"
DATABASE_PATH = ROOT / "waternet" / "data" / "waternet.db"
SEGMENT_PROFILE_XLSX = Path(
    "D:/下载wechat/xwechat_files/wxid_q0cjhyijiy6f12_731b/temp/RWTemp/2026-07/"
    "b3529b0272f03e3041b1ed889b154a58/72小时河段水面线流量分布.xlsx"
)
NODE_PROCESS_XLSX = Path(
    "D:/下载wechat/xwechat_files/wxid_q0cjhyijiy6f12_731b/temp/RWTemp/2026-07/"
    "b3529b0272f03e3041b1ed889b154a58/72小时全节点水位流量过程.xlsx"
)


def main():
    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DATABASE_PATH.exists():
        DATABASE_PATH.unlink()

    with sqlite3.connect(DATABASE_PATH) as connection:
        connection.execute("pragma foreign_keys = on")
        create_schema(connection)
        seed_users(connection)
        seed_platform_observations(connection)
        seed_nodes(connection)
        seed_segments(connection)
        seed_segment_profiles(connection)
        seed_node_timeseries(connection)
        refresh_derived_tables(connection)

    print(f"created {DATABASE_PATH}")


def create_schema(connection):
    connection.executescript(
        """
        create table users (
          user_id integer primary key autoincrement,
          username text not null unique,
          password text not null,
          role text not null,
          account_status text not null check (account_status in ('ENABLED', 'DISABLED')),
          created_at text not null,
          last_login_at text
        );

        create table river_segments (
          code text primary key,
          name text not null,
          reach_id integer not null unique,
          start_node_code text,
          end_node_code text,
          length_meters real not null,
          width_meters real not null,
          dx real,
          chezy real,
          bed_elevation real,
          coordinates_json text not null
        );

        create table water_nodes (
          code text primary key,
          node_id integer not null unique,
          name text not null,
          node_type text not null,
          lng real not null,
          lat real not null,
          initial_water_level real not null,
          boundary_type text not null,
          connected_node_codes_json text not null,
          connected_segment_codes_json text not null,
          connected_reach_ids_json text not null
        );

        create table platform_observations (
          id integer primary key autoincrement,
          observed_at text not null,
          taihu_water_level real,
          canal_north_flow real,
          canal_south_water_level real
        );

        create table segment_profile_results (
          id integer primary key autoincrement,
          segment_code text not null,
          reach_id integer not null,
          start_node_code text,
          end_node_code text,
          section_no integer not null,
          x_m real not null,
          water_level real not null,
          flow real not null
        );

        create table node_hydrology_timeseries (
          id integer primary key autoincrement,
          hour integer not null,
          node_code text not null,
          water_level real,
          flow real,
          unique(hour, node_code)
        );

        create table segment_hydrology_stats (
          segment_code text primary key,
          max_flow real,
          min_flow real,
          max_water_level real,
          min_water_level real,
          profile_hour integer,
          sample_count integer
        );

        create table node_latest_hydrology (
          node_code text primary key,
          hour integer,
          water_level real,
          flow real
        );
        """
    )


def seed_users(connection):
    connection.executemany(
        """
        insert into users (username, password, role, account_status, created_at, last_login_at)
        values (?, ?, ?, ?, datetime('now'), null)
        """,
        [
            ("admin", "{noop}admin123", "ADMIN", "ENABLED"),
            ("operator", "{noop}operator123", "OPERATOR", "ENABLED"),
        ],
    )


def seed_platform_observations(connection):
    connection.execute(
        """
        insert into platform_observations
          (observed_at, taihu_water_level, canal_north_flow, canal_south_water_level)
        values (datetime('now'), 13.2, -213.3138, 13.3)
        """
    )


def seed_nodes(connection):
    nodes_geojson = read_json(FRONTEND_DATA / "water-nodes.geojson")
    topology = read_json(FRONTEND_DATA / "node-topology.json")
    rows = []
    for feature in nodes_geojson["features"]:
        node_id = int(feature["properties"]["name"])
        code = node_code(node_id)
        category = classify_node(node_id)
        lng, lat = feature["geometry"]["coordinates"]
        topo = topology.get(code, {})
        rows.append(
            (
                code,
                node_id,
                f"{category['label']} {node_id}",
                category["type"],
                float(lng),
                float(lat),
                2.44,
                category["boundaryType"],
                json.dumps(topo.get("connectedNodeCodes", []), ensure_ascii=False),
                json.dumps(topo.get("connectedSegmentCodes", []), ensure_ascii=False),
                json.dumps(topo.get("connectedReachIds", []), ensure_ascii=False),
            )
        )

    connection.executemany(
        """
        insert into water_nodes (
          code, node_id, name, node_type, lng, lat, initial_water_level, boundary_type,
          connected_node_codes_json, connected_segment_codes_json, connected_reach_ids_json
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )


def seed_segments(connection):
    segment_geojson = read_json(FRONTEND_DATA / "river-segments-67.geojson")
    topology = read_json(FRONTEND_DATA / "node-topology.json")
    metadata_by_reach = read_segment_metadata()
    endpoints_by_reach = endpoints_from_topology(topology)
    rows = []

    for index, feature in enumerate(segment_geojson["features"], start=1):
        reach_id = int(feature.get("properties", {}).get("reachId") or feature.get("properties", {}).get("fid") or index)
        code = f"REAL_RIVER_{reach_id:02d}"
        metadata = metadata_by_reach.get(reach_id, {})
        endpoints = endpoints_by_reach.get(reach_id, {})
        coordinates = normalize_line_coordinates(feature["geometry"])
        rows.append(
            (
                code,
                metadata.get("name") or f"河段 {reach_id}",
                reach_id,
                metadata.get("startNodeCode") or endpoints.get("start"),
                metadata.get("endNodeCode") or endpoints.get("end"),
                float(metadata.get("length") or calculate_length_meters(coordinates)),
                float(metadata.get("width") or 24),
                metadata.get("dx"),
                metadata.get("chezy"),
                metadata.get("bed"),
                json.dumps([{"lng": lng, "lat": lat} for lng, lat in coordinates], ensure_ascii=False),
            )
        )

    connection.executemany(
        """
        insert into river_segments (
          code, name, reach_id, start_node_code, end_node_code, length_meters, width_meters,
          dx, chezy, bed_elevation, coordinates_json
        ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )


def seed_segment_profiles(connection):
    workbook = openpyxl.load_workbook(SEGMENT_PROFILE_XLSX, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        reach_id = int(row[0])
        rows.append(
            (
                f"REAL_RIVER_{reach_id:02d}",
                reach_id,
                node_code(row[1]) if row[1] is not None else None,
                node_code(row[2]) if row[2] is not None else None,
                int(row[3]),
                float(row[4]),
                float(row[5]),
                float(row[6]),
            )
        )

    connection.executemany(
        """
        insert into segment_profile_results (
          segment_code, reach_id, start_node_code, end_node_code, section_no, x_m, water_level, flow
        ) values (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )


def seed_node_timeseries(connection):
    workbook = openpyxl.load_workbook(NODE_PROCESS_XLSX, read_only=True, data_only=True)
    level_sheet = workbook["节点水位"]
    flow_sheet = workbook["节点流量"]
    level_rows = list(level_sheet.iter_rows(values_only=True))
    flow_rows = list(flow_sheet.iter_rows(values_only=True))
    headers = level_rows[0]
    node_ids = [parse_node_id(header) for header in headers[1:]]
    flow_by_hour = {int(row[0]): row for row in flow_rows[1:] if row[0] is not None}
    rows = []

    for level_row in level_rows[1:]:
        if level_row[0] is None:
            continue
        hour = int(level_row[0])
        flow_row = flow_by_hour.get(hour)
        for index, node_id in enumerate(node_ids, start=1):
            if node_id is None:
                continue
            water_level = level_row[index] if index < len(level_row) else None
            flow = flow_row[index] if flow_row and index < len(flow_row) else None
            rows.append(
                (
                    hour,
                    node_code(node_id),
                    float(water_level) if water_level is not None else None,
                    float(flow) if flow is not None else None,
                )
            )

    connection.executemany(
        """
        insert into node_hydrology_timeseries (hour, node_code, water_level, flow)
        values (?, ?, ?, ?)
        """,
        rows,
    )


def refresh_derived_tables(connection):
    connection.executescript(
        """
        insert into segment_hydrology_stats (
          segment_code, max_flow, min_flow, max_water_level, min_water_level, profile_hour, sample_count
        )
        select
          segment_code,
          max(flow),
          min(flow),
          max(water_level),
          min(water_level),
          72,
          count(*)
        from segment_profile_results
        group by segment_code;

        insert into node_latest_hydrology (node_code, hour, water_level, flow)
        select node_code, hour, water_level, flow
        from node_hydrology_timeseries
        where hour = (select max(hour) from node_hydrology_timeseries);
        """
    )


def read_segment_metadata():
    path = FRONTEND_DATA / "river-network.geojson"
    if not path.exists():
        return {}
    data = read_json(path)
    result = {}
    for feature in data.get("features", []):
        properties = feature.get("properties", {})
        reach_id = properties.get("reachId") or properties.get("fid")
        if reach_id is not None:
            result[int(reach_id)] = properties
    return result


def endpoints_from_topology(topology):
    result = {}
    for code, record in topology.items():
        for reach_id in record.get("connectedReachIds", []):
            item = result.setdefault(int(reach_id), [])
            item.append(code)
    return {
        reach_id: {"start": codes[0], "end": codes[-1]}
        for reach_id, codes in result.items()
        if codes
    }


def normalize_line_coordinates(geometry):
    coordinates = geometry["coordinates"]
    if geometry["type"] == "LineString":
        pairs = coordinates
    else:
        pairs = [pair for line in coordinates for pair in line]
    if any(abs(pair[0]) > 180 or abs(pair[1]) > 90 for pair in pairs):
        return [web_mercator_to_lng_lat(x, y) for x, y in pairs]
    return [(float(x), float(y)) for x, y in pairs]


def web_mercator_to_lng_lat(x, y):
    import math

    radius = 6378137
    lng = (float(x) / radius) * (180 / math.pi)
    lat = (2 * math.atan(math.exp(float(y) / radius)) - math.pi / 2) * (180 / math.pi)
    return (lng, lat)


def calculate_length_meters(coordinates):
    return sum(distance_meters(coordinates[index - 1], coordinates[index]) for index in range(1, len(coordinates)))


def distance_meters(left, right):
    import math

    earth_radius = 6371000
    lat1 = math.radians(left[1])
    lat2 = math.radians(right[1])
    delta_lat = math.radians(right[1] - left[1])
    delta_lng = math.radians(right[0] - left[0])
    a = math.sin(delta_lat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(delta_lng / 2) ** 2
    return earth_radius * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def parse_node_id(value):
    if value is None:
        return None
    match = re.search(r"节点(\d+)#", str(value))
    return int(match.group(1)) if match else None


def node_code(value):
    return f"N{int(value):02d}"


def classify_node(node_id):
    if node_id in (1, 3, 6):
        return {"label": "边界节点", "type": "BOUNDARY", "boundaryType": "BOUNDARY"}
    if 7 <= node_id <= 20:
        return {"label": "断头河节点", "type": "DEAD_END", "boundaryType": "DEAD_END"}
    return {"label": "节点", "type": "JUNCTION", "boundaryType": "NONE"}


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8-sig"))


if __name__ == "__main__":
    main()
